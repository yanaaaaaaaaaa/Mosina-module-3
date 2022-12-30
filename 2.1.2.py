import csv
import math
from time import sleep

import matplotlib
import tkinter
import openpyxl.styles.numbers
import openpyxl
import matplotlib.pyplot as plt


class Vacancy:
    def __init__(self, **kwargs):
        self.salary_currency = kwargs['salary_currency']
        self.area_name = kwargs['area_name']
        self.name = kwargs['name']
        self.__set_published_at(kwargs['published_at'])
        self.salary = (float(kwargs['salary_from']) + float(kwargs['salary_to'])) // 2

    def __set_published_at(self, published_at):
        spliten = published_at.split('T')
        date = spliten[0].split('-')
        self.year = int(date[0])


class DataSet:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055}

    def __init__(self, header):
        self.header = header
        self.vacancies: list[Vacancy] = []

    def print_stats(self, vacancy_name, stats=None):
        if stats is None:
            stats = self.get_stats(vacancy_name)

        print('Динамика уровня зарплат по годам:', stats['salary_stat'])
        print('Динамика количества вакансий по годам:', stats['vacancy_count_stat'])
        print('Динамика уровня зарплат по годам для выбранной профессии:', stats['selected_salary_stat'])
        print('Динамика количества вакансий по годам для выбранной профессии:', stats['selected_count_stat'])
        print('Уровень зарплат по городам (в порядке убывания):',
              {k: stats['area_salary_stat'][k] for i, k in zip(range(10), stats['area_salary_stat'])})
        print('Доля вакансий по городам (в порядке убывания):', {k: stats['doly_stat'][k] for i, k in zip(range(10), stats['doly_stat'])})

    def get_stats(self, vacancy_name):
        vacancies = self.vacancies
        doly_stat = dict()
        salary_stat = dict()
        vacancy_count_stat = dict()
        selected_salary_stat = dict()
        selected_count_stat = dict()
        area_salary_stat = dict()
        area_count_stat = dict()
        for vacancy in vacancies:
            salary = int(vacancy.salary * DataSet.currency_to_rub[vacancy.salary_currency])
            if vacancy.area_name not in doly_stat:
                doly_stat[vacancy.area_name] = 0
                area_salary_stat[vacancy.area_name] = 0
                area_count_stat[vacancy.area_name] = 0
            doly_stat[vacancy.area_name] += 1
            area_salary_stat[vacancy.area_name] += salary
            area_count_stat[vacancy.area_name] += 1

            if vacancy.year not in salary_stat:
                salary_stat[vacancy.year] = 0
                vacancy_count_stat[vacancy.year] = 0
                selected_salary_stat[vacancy.year] = 0
                selected_count_stat[vacancy.year] = 0
            salary_stat[vacancy.year] += salary
            vacancy_count_stat[vacancy.year] += 1
            if vacancy_name in vacancy.name:
                selected_salary_stat[vacancy.year] += salary
                selected_count_stat[vacancy.year] += 1

        doly_stat = {k: doly_stat[k] / len(vacancies) for k in doly_stat if doly_stat[k] >= int(len(vacancies) / 100)}
        doly_stat = {k: round(doly_stat[k], 4) for k in sorted(doly_stat, key=lambda k: -doly_stat[k])}
        salary_stat = {k: salary_stat[k] // vacancy_count_stat[k] for k in sorted(salary_stat)}
        selected_salary_stat = {k: selected_salary_stat[k] // selected_count_stat[k] if selected_count_stat[k] != 0 else 0 for k in sorted(selected_salary_stat)}
        area_salary_stat = {k: area_salary_stat[k] // area_count_stat[k] for k in area_salary_stat if k in doly_stat}
        area_salary_stat = {k: area_salary_stat[k] for k in sorted(area_salary_stat, key=lambda k: -area_salary_stat[k])}
        return {'doly_stat': doly_stat, 'vacancy_count_stat': vacancy_count_stat, 'salary_stat': salary_stat, 'area_salary_stat': area_salary_stat, 'selected_salary_stat': selected_salary_stat, 'selected_count_stat': selected_count_stat}


class report:
    def __init__(self, **kwargs):
        self._wb = openpyxl.Workbook()
        self._header_font = kwargs['header_font'] if 'header_font' in kwargs else None
        self._font = kwargs['font'] if 'font' in kwargs else None
        self._border = kwargs['border'] if 'border' in kwargs else None

    def generate_excel(self, stats, vacancy_name):
        ws = self._wb.active
        ws.title = 'Статистика по годам'
        header = ['Год', 'Средняя зарплата', f'Средняя зарплата - {vacancy_name}', 'Количество вакансий',
                  f'Количество вакансий - {vacancy_name}']
        ws.append(header)
        for year in stats[0]:
            row = [year]
            row.extend(d[year] for d in stats if year in d)
            ws.append(row)
        self.__formating_sheet(len(header), len(stats[0]), header, ws)

        ws = self._wb.create_sheet('Статистика по городам')
        header = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']
        ws.append(header)
        for city, _ in zip(stats[-1], range(10)):
            row = [city]
            row.append(stats[-2][city])
            row.append('')
            row.append(city)
            row.append(stats[-1][city])
            ws.append(row)
        for row in ws.iter_rows(min_row=2, max_row=11, min_col=5):
            for cell in row:
                cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]
        self.__formating_sheet(len(header), 10, header, ws)
        self._wb.save('report.xlsx')

    def generate_image(self, stats, vacancy_name):
        matplotlib.use('TkAgg')
        plt.rc('font', size=8)
        plt.rc('axes', titlesize=13)

        figure = plt.figure()
        ax1 = figure.add_subplot(2, 2, 1)
        x = list(stats['salary_stat'])
        y1 = [stats['salary_stat'][k] for k in x]
        y2 = [stats['selected_salary_stat'][k] for k in x]
        ax1.bar([i - 0.2 for i in x], y1, label='средняя з/п', width=0.4)
        ax1.bar([i + 0.2 for i in x], y2, label=f'з/п {vacancy_name}', width=0.4)
        ax1.set_xticks(x, fontsize=8)
        ax1.tick_params(axis='x', labelrotation=90)
        ax1.grid(axis='y')
        ax1.set(title='Уровень зарплат по годам')

        ax2 = figure.add_subplot(2, 2, 2)
        x = list(stats['vacancy_count_stat'])
        y1 = [stats['vacancy_count_stat'][k] for k in x]
        y2 = [stats['selected_count_stat'][k] for k in x]
        ax2.bar([i - 0.2 for i in x], y1, label='Количество вакансий', width=0.4)
        ax2.bar([i + 0.2 for i in x], y2, label=f'Количество вакансий {vacancy_name}', width=0.4)
        ax2.set_xticks(x, fontsize=8)
        ax2.tick_params(axis='x', labelrotation=90)
        ax2.grid(axis='y')
        ax2.set(title='Количество вакансий по годам')

        ax3 = figure.add_subplot(2, 2, 3)
        x = list(stats['area_salary_stat'])[:10]
        y = [stats['area_salary_stat'][k] for k in x]
        x = [i if ' ' not in i and '-' not in i else i.replace(' ', '\n').replace('-', '-\n') for i in x]
        ax3.barh(x, y)
        ax3.grid(axis='x')
        ax3.set(title='Уровень зарплат по городам')

        ax4 = figure.add_subplot(2, 2, 4)
        x = list(stats['doly_stat'])[:10][::-1]
        y = [stats['doly_stat'][k] for k in x]
        x.append('Другие')
        y.append(1 - sum(y))
        ax4.pie(y, labels=x)
        ax4.set(title='Доля вакансий по городам')

        plt.subplots_adjust(wspace=0.5, hspace=0.5)
        plt.show()
        figure.savefig('graph.png')


    def __formating_sheet(self, row_len, column_len, header, sheet):
        A_ord = ord('A')
        for column_ord in range(row_len):
            column = chr(column_ord + A_ord)
            sheet.column_dimensions[column].width = len(header[column_ord]) + 2
        for row in sheet.iter_rows(max_row=1, max_col=row_len):
            for cell in row:
                cell.font = self._header_font
                cell.border = self._border
        for row in sheet.iter_rows(max_row=column_len+1, min_row=2, max_col=row_len):
            for cell in row:
                cell.font = self._font
                cell.border = self._border


def read_csv(filename):
    with open(filename, encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        header = next(reader)
        data_set = DataSet(header)
        data_set.vacancies = [Vacancy(**{k: v for k, v in zip(header, line)}) for line in reader if len(line) == len(header) and all(line)]
    return data_set


file_name = input('Введите название файла: ')
vacancy_name = input('Введите название профессии: ')
data_set = read_csv(file_name)
stats = data_set.get_stats(vacancy_name)
data_set.print_stats(vacancy_name, stats)

report = report()
report.generate_image(stats, vacancy_name)

